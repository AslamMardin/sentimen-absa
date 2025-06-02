import pandas as pd
import numpy as np
from transformers import T5ForConditionalGeneration, T5Tokenizer, AutoTokenizer, AutoModelForSeq2SeqLM
import torch
from tqdm import tqdm
import re
import time
import random
import string
from collections import Counter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Untuk LDA
import gensim
from gensim import corpora
from gensim.models import LdaModel
from gensim.utils import simple_preprocess
import nltk
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize

# Download NLTK data jika belum ada
try:
    nltk.data.find('tokenizers/punkt')
except LookupError:
    nltk.download('punkt')

try:
    nltk.data.find('corpora/stopwords')
except LookupError:
    nltk.download('stopwords')

class CompleteTextProcessor:
    def __init__(self, paraphrase_model="cahya/bloomz-1b7-instruct"):
        """Inisialisasi semua komponen text processing"""
        
        print("🚀 Menginisialisasi Complete Text Processor...")
        
        # Inisialisasi Paraphraser
        self.init_paraphraser(paraphrase_model)
        
        # Inisialisasi LDA
        self.aspect_labels = {
            0: "Kualitas Guru",
            1: "Fasilitas", 
            2: "Lingkungan",
            3: "Kegiatan Pondok",
            4: "Pembinaan Karakter",
            5: "Prestasi",
            6: "Akademik",
            7: "Motivasi/Spiritual",
            8: "Sosial",
            9: "Umum"
        }
        
        # Stopwords Indonesia
        self.indonesian_stopwords = set([
            'yang', 'di', 'ke', 'dari', 'dalam', 'untuk', 'pada', 'dengan', 'oleh', 
            'karena', 'sebagai', 'adalah', 'akan', 'telah', 'sudah', 'masih', 'dapat',
            'bisa', 'juga', 'hanya', 'atau', 'dan', 'ini', 'itu', 'saya', 'kami',
            'kita', 'mereka', 'dia', 'ia', 'nya', 'mu', 'ku', 'se', 'ter', 'ber',
            'me', 'pe', 'ke', 'hal', 'cara', 'banyak', 'satu', 'dua', 'tiga',
            'tidak', 'bukan', 'belum', 'jangan', 'agar', 'supaya', 'kalau', 'jika'
        ])
        
        # Sentiment lexicons (akan dimuat dari file)
        self.positive_words = set()
        self.negative_words = set()
        
        print("✅ Complete Text Processor siap digunakan!")
    
    def init_paraphraser(self, model_name):
        """Inisialisasi model parafrase"""
        print(f"📚 Loading paraphrase model: {model_name}")
        self.device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
        print(f"💻 Using device: {self.device}")
        
        try:
            self.tokenizer = AutoTokenizer.from_pretrained(model_name)
            self.model = AutoModelForSeq2SeqLM.from_pretrained(model_name)
            self.model.to(self.device)
            self.model.eval()
            print("✅ Paraphrase model berhasil dimuat!")
        except Exception as e:
            print(f"❌ Error loading model: {e}")
            print("🔄 Mencoba model alternatif...")
            model_name = "cahya/bloomz-1b7-instruct"
            self.tokenizer = AutoTokenizer.from_pretrained(model_name)
            self.model = AutoModelForSeq2SeqLM.from_pretrained(model_name)
            self.model.to(self.device)
            self.model.eval()
            print("✅ Model alternatif berhasil dimuat!")
    
    def load_sentiment_lexicons(self, positive_file="positive.csv", negative_file="negative.csv"):
        """Load kata-kata positif dan negatif dari file CSV"""
        print("📖 Loading sentiment lexicons...")
        
        try:
            # Load positive words - support both CSV and Excel
            if positive_file.endswith('.xlsx') or positive_file.endswith('.xls'):
                pos_df = pd.read_excel(positive_file)
            else:
                pos_df = pd.read_csv(positive_file)
                
            if 'word' in pos_df.columns:
                self.positive_words = set(pos_df['word'].astype(str).str.lower())
            elif len(pos_df.columns) > 0:
                # Ambil kolom pertama jika tidak ada kolom 'word'
                self.positive_words = set(pos_df.iloc[:, 0].astype(str).str.lower())
            
            print(f"✅ Loaded {len(self.positive_words)} positive words")
            
        except Exception as e:
            print(f"⚠️ Warning loading positive words: {e}")
            # Default positive words Indonesia
            self.positive_words = set([
                'baik', 'bagus', 'hebat', 'luar biasa', 'sempurna', 'excellent', 'mantap',
                'senang', 'puas', 'suka', 'cinta', 'indah', 'cantik', 'keren', 'oke',
                'positif', 'optimis', 'berhasil', 'sukses', 'juara', 'terbaik', 'unggul'
            ])
        
        try:
            # Load negative words - support both CSV and Excel
            if negative_file.endswith('.xlsx') or negative_file.endswith('.xls'):
                neg_df = pd.read_excel(negative_file)
            else:
                neg_df = pd.read_csv(negative_file)
                
            if 'word' in neg_df.columns:
                self.negative_words = set(neg_df['word'].astype(str).str.lower())
            elif len(neg_df.columns) > 0:
                self.negative_words = set(neg_df.iloc[:, 0].astype(str).str.lower())
            
            print(f"✅ Loaded {len(self.negative_words)} negative words")
            
        except Exception as e:
            print(f"⚠️ Warning loading negative words: {e}")
            # Default negative words Indonesia
            self.negative_words = set([
                'buruk', 'jelek', 'tidak baik', 'kecewa', 'sedih', 'marah', 'benci',
                'gagal', 'kalah', 'lemah', 'bodoh', 'rusak', 'kotor', 'jorok',
                'negatif', 'pesimis', 'susah', 'sulit', 'parah', 'terburuk'
            ])
    
    def clean_text_for_processing(self, text):
        """Membersihkan teks untuk semua jenis processing"""
        if pd.isna(text) or text == "":
            return ""
        
        text = str(text).lower()
        # Hapus tanda baca dan karakter khusus
        text = re.sub(r'[^\w\s]', ' ', text)
        # Hapus multiple spaces
        text = re.sub(r'\s+', ' ', text)
        text = text.strip()
        
        return text
    
    def preprocess_for_lda(self, texts):
        """Preprocessing khusus untuk LDA"""
        processed_texts = []
        
        for text in tqdm(texts, desc="Preprocessing for LDA"):
            if pd.isna(text) or text == "":
                processed_texts.append([])
                continue
                
            # Clean text
            clean_text = self.clean_text_for_processing(text)
            
            # Tokenize
            tokens = simple_preprocess(clean_text, deacc=True)
            
            # Remove stopwords
            tokens = [token for token in tokens if token not in self.indonesian_stopwords 
                     and len(token) > 2]
            
            processed_texts.append(tokens)
        
        return processed_texts
    
    def train_lda_model(self, texts, num_topics=10, random_state=42):
        """Training LDA model"""
        print("🎯 Training LDA model untuk aspect classification...")
        
        # Preprocess texts
        processed_texts = self.preprocess_for_lda(texts)
        
        # Filter empty documents
        processed_texts = [text for text in processed_texts if len(text) > 0]
        
        if len(processed_texts) == 0:
            print("❌ No valid documents for LDA training!")
            return None, None
        
        # Create dictionary dan corpus
        dictionary = corpora.Dictionary(processed_texts)
        dictionary.filter_extremes(no_below=2, no_above=0.8)
        corpus = [dictionary.doc2bow(text) for text in processed_texts]
        
        # Train LDA model
        lda_model = LdaModel(
            corpus=corpus,
            id2word=dictionary,
            num_topics=num_topics,
            random_state=42,
            passes=10,
            alpha='auto'
        )
        
        print("✅ LDA model berhasil dilatih!")
        
        # Print topics untuk referensi
        print("\n📋 Topics yang ditemukan:")
        for idx, topic in lda_model.print_topics(-1):
            print(f"Topic {idx}: {topic}")
        
        return lda_model, dictionary
    
    def predict_aspect(self, text, lda_model, dictionary):
        """Prediksi aspek menggunakan LDA model"""
        if pd.isna(text) or text == "" or lda_model is None:
            return 9  # Default ke "Umum"
        
        # Preprocess text
        clean_text = self.clean_text_for_processing(text)
        tokens = simple_preprocess(clean_text, deacc=True)
        tokens = [token for token in tokens if token not in self.indonesian_stopwords 
                 and len(token) > 2]
        
        if len(tokens) == 0:
            return 9
        
        # Create bow representation
        bow = dictionary.doc2bow(tokens)
        
        if len(bow) == 0:
            return 9
        
        # Get topic distribution
        topic_dist = lda_model[bow]
        
        if len(topic_dist) == 0:
            return 9
        
        # Return topic with highest probability
        best_topic = max(topic_dist, key=lambda x: x[1])
        return best_topic[0]
    
    def analyze_sentiment(self, text):
        """Analisis sentimen menggunakan lexicon-based approach"""
        if pd.isna(text) or text == "":
            return "netral"
        
        # Clean and tokenize
        clean_text = self.clean_text_for_processing(text)
        words = clean_text.split()
        
        # Count positive and negative words
        positive_count = sum(1 for word in words if word in self.positive_words)
        negative_count = sum(1 for word in words if word in self.negative_words)
        
        # Simple rule-based classification
        if positive_count > negative_count:
            return "positif"
        elif negative_count > positive_count:
            return "negatif"
        else:
            return "netral"
    
    def paraphrase_text(self, text, temperature=1.1):
        """Parafrase teks Indonesia"""
        if pd.isna(text) or text == "" or len(str(text).strip()) < 10:
            return text
        
        clean_text = str(text).strip()
        if len(clean_text) > 400:
            clean_text = clean_text[:400] + "..."
        
        try:
            input_text = f"parafrase: {clean_text}"
            
            inputs = self.tokenizer(
                input_text,
                return_tensors="pt",
                max_length=512,
                truncation=True,
                padding=True
            ).to(self.device)
            
            with torch.no_grad():
                outputs = self.model.generate(
                    **inputs,
                    max_length=256,
                    min_length=len(clean_text.split()) // 2,
                    num_beams=4,
                    temperature=temperature,
                    do_sample=True,
                    repetition_penalty=1.2,
                    length_penalty=1.0,
                    early_stopping=True,
                    pad_token_id=self.tokenizer.pad_token_id,
                    eos_token_id=self.tokenizer.eos_token_id
                )
            
            paraphrased = self.tokenizer.decode(outputs[0], skip_special_tokens=True)
            paraphrased = re.sub(r'^parafrase:\s*', '', paraphrased, flags=re.IGNORECASE)
            paraphrased = paraphrased.strip()
            
            if paraphrased and not paraphrased[0].isupper():
                paraphrased = paraphrased[0].upper() + paraphrased[1:]
            
            return paraphrased if len(paraphrased) > len(clean_text) * 0.5 else text
            
        except Exception as e:
            print(f"Error paraphrasing: {e}")
            return text

def save_to_excel_with_formatting(df, output_file):
    """Simpan DataFrame ke Excel dengan formatting yang rapi"""
    print(f"💾 Menyimpan ke Excel: {output_file}")
    
    # Simpan ke Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Processed_Data', index=False)
        
        # Akses workbook dan worksheet untuk formatting
        workbook = writer.book
        worksheet = writer.sheets['Processed_Data']
        
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Format header row
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Max width 50
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Color coding untuk sentiment
        for row_num in range(2, len(df) + 2):  # Skip header
            sentiment_col = None
            for col_num, col_name in enumerate(df.columns, 1):
                if col_name == 'predicted_sentiment':
                    sentiment_col = col_num
                    break
            
            if sentiment_col:
                cell = worksheet.cell(row=row_num, column=sentiment_col)
                if cell.value == 'positif':
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif cell.value == 'negatif':
                    cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                elif cell.value == 'netral':
                    cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
        
        # Freeze panes pada header
        worksheet.freeze_panes = "A2"
    
    print("✅ File Excel berhasil disimpan dengan formatting!")

def process_complete_excel(input_file, output_file, text_column='text_combined', 
                          refined_column='text_refined', target_count=800,
                          positive_file="positive.csv", negative_file="negative.csv"):
    """
    Fungsi utama untuk memproses Excel lengkap dengan parafrase, LDA, dan sentiment analysis
    """
    
    print("🚀 Memulai Complete Text Processing untuk Excel...")
    
    # Load Excel
    try:
        df = pd.read_excel(input_file)
        print(f"✅ File Excel berhasil dimuat: {len(df)} baris")
        print(f"📋 Kolom yang tersedia: {list(df.columns)}")
    except Exception as e:
        print(f"❌ Error membaca Excel: {e}")
        return
    
    # Cek kolom yang dibutuhkan
    required_columns = [text_column]
    if refined_column not in df.columns:
        print(f"⚠️ Kolom '{refined_column}' tidak ditemukan, akan menggunakan '{text_column}'")
        refined_column = text_column
    
    for col in required_columns:
        if col not in df.columns:
            print(f"❌ Kolom '{col}' tidak ditemukan!")
            print(f"Kolom yang tersedia: {list(df.columns)}")
            return
    
    # Inisialisasi processor
    processor = CompleteTextProcessor()
    
    # Load sentiment lexicons
    processor.load_sentiment_lexicons(positive_file, negative_file)
    
    # Train LDA model pada data asli
    print("\n🎯 Training LDA model...")
    lda_model, dictionary = processor.train_lda_model(df[text_column].fillna(""), num_topics=10)
    
    # Proses data asli dulu
    print("\n📊 Memproses data asli...")
    original_data = df.copy()
    
    # Prediksi aspek dan sentimen untuk data asli
    tqdm.pandas(desc="Predicting aspects")
    original_data['predicted_aspect'] = original_data[text_column].progress_apply(
        lambda x: processor.predict_aspect(x, lda_model, dictionary)
    )
    
    tqdm.pandas(desc="Analyzing sentiment")
    original_data['predicted_sentiment'] = original_data[refined_column].progress_apply(
        processor.analyze_sentiment
    )
    
    # Tambahkan label aspek yang readable
    original_data['aspect_label'] = original_data['predicted_aspect'].map(processor.aspect_labels)
    
    # Siapkan data untuk parafrase
    paraphrased_data = []
    current_count = len(original_data)
    iterations_needed = max(1, target_count // current_count)
    
    print(f"\n📈 Data saat ini: {current_count} baris")
    print(f"🎯 Target: {target_count} baris")
    print(f"🔄 Akan melakukan {iterations_needed} iterasi parafrase")
    
    # Lakukan parafrase
    for iteration in range(iterations_needed):
        print(f"\n🔄 Iterasi parafrase {iteration + 1}/{iterations_needed}")
        
        df_iteration = original_data.sample(frac=1).reset_index(drop=True)
        
        for idx, row in tqdm(df_iteration.iterrows(), 
                           total=len(df_iteration), 
                           desc=f"Parafrase Iterasi {iteration + 1}"):
            
            # Parafrase teks
            original_text = row[text_column]
            paraphrased_text = processor.paraphrase_text(
                original_text, 
                temperature=random.uniform(1.0, 1.3)
            )
            
            # Buat row baru
            new_row = row.copy()
            new_row[text_column] = paraphrased_text
            
            # Prediksi aspek dan sentimen untuk teks yang diparafrase
            new_row['predicted_aspect'] = processor.predict_aspect(
                paraphrased_text, lda_model, dictionary
            )
            new_row['predicted_sentiment'] = processor.analyze_sentiment(
                paraphrased_text if refined_column == text_column else row[refined_column]
            )
            new_row['aspect_label'] = processor.aspect_labels[new_row['predicted_aspect']]
            
            paraphrased_data.append(new_row)
            
            if len(paraphrased_data) >= target_count - current_count:
                break
        
        if len(paraphrased_data) >= target_count - current_count:
            break
    
    # Gabungkan semua data
    final_df = pd.concat([original_data, pd.DataFrame(paraphrased_data)], 
                        ignore_index=True)
    
    # Batasi sesuai target
    if len(final_df) > target_count:
        final_df = final_df.sample(n=target_count).reset_index(drop=True)
    
    # Simpan hasil
    try:
        save_to_excel_with_formatting(final_df, output_file)
        print(f"\n✅ Berhasil menyimpan {len(final_df)} baris ke {output_file}")
        
        # Statistik hasil
        print("\n📊 STATISTIK HASIL:")
        print(f"📈 Total data: {len(final_df)}")
        print(f"📈 Data asli: {current_count}")
        print(f"📈 Data parafrase: {len(paraphrased_data)}")
        
        print("\n🎯 DISTRIBUSI ASPEK:")
        aspect_dist = final_df['aspect_label'].value_counts()
        for aspect, count in aspect_dist.items():
            print(f"   {aspect}: {count} ({count/len(final_df)*100:.1f}%)")
        
        print("\n😊 DISTRIBUSI SENTIMEN:")
        sentiment_dist = final_df['predicted_sentiment'].value_counts()
        for sentiment, count in sentiment_dist.items():
            print(f"   {sentiment}: {count} ({count/len(final_df)*100:.1f}%)")
        
    except Exception as e:
        print(f"❌ Error menyimpan file: {e}")

# Contoh penggunaan
if __name__ == "__main__":
    # Konfigurasi
    INPUT_FILE = "hasil_sentimen_pesantren.xlsx"  # File Excel input
    OUTPUT_FILE = "complete_processed_pesantren.xlsx"  # File Excel output
    TEXT_COLUMN = "text_combined"  # Kolom teks utama
    REFINED_COLUMN = "text_refined"  # Kolom teks untuk sentiment analysis
    TARGET_COUNT = 800  # Target jumlah data
    POSITIVE_FILE = "positive.csv"  # File kata positif (bisa .csv atau .xlsx)
    NEGATIVE_FILE = "negative.csv"  # File kata negatif (bisa .csv atau .xlsx)
    
    # Jalankan processing lengkap
    process_complete_excel(
        input_file=INPUT_FILE,
        output_file=OUTPUT_FILE,
        text_column=TEXT_COLUMN,
        refined_column=REFINED_COLUMN,
        target_count=TARGET_COUNT,
        positive_file=POSITIVE_FILE,
        negative_file=NEGATIVE_FILE
    )
    
    print("\n🎉 Complete Text Processing selesai!")
    print("💡 File output berisi kolom tambahan:")
    print("   - predicted_aspect: Prediksi aspek (0-9)")
    print("   - aspect_label: Label aspek yang readable")
    print("   - predicted_sentiment: Sentimen (positif/negatif/netral)")